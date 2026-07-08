"""
Sprint 3

Peer Comparison Report

Generates:
output/peer_comparison.xlsx

One worksheet per peer group using the
latest available financial year.
"""

from pathlib import Path
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ==========================================================
# Paths
# ==========================================================

BASE_DIR = Path(__file__).resolve().parents[2]

DB_PATH = BASE_DIR / "db" / "nifty100.db"

OUTPUT_PATH = BASE_DIR / "output" / "peer_comparison.xlsx"


# ==========================================================
# Excel Colours
# ==========================================================

GREEN = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)

YELLOW = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)

RED = PatternFill(
    start_color="F4CCCC",
    end_color="F4CCCC",
    fill_type="solid",
)

GOLD = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid",
)


# ==========================================================
# Year Ordering
# ==========================================================

YEAR_PRIORITY = {
    "Dec 2012": 2012,
    "Mar 2013": 2013,
    "Mar 2014": 2014,
    "Mar 2015": 2015,
    "Mar 2016": 2016,
    "Mar 2017": 2017,
    "Mar 2018": 2018,
    "Mar 2019": 2019,
    "Mar 2020": 2020,
    "Mar 2021": 2021,
    "Mar 2022": 2022,
    "Mar 2023": 2023,
    "Mar 2024": 2024,
    "Sep 2024": 2024.5,
    "TTM": 2025,
}


# ==========================================================
# Metrics
# ==========================================================

RANK_METRICS = {
    "return_on_equity_pct": False,
    "return_on_capital_employed_pct": False,
    "net_profit_margin_pct": False,
    "debt_to_equity": True,
    "free_cash_flow_cr": False,
    "pat_cagr_5yr": False,
    "revenue_cagr_5yr": False,
    "interest_coverage": False,
    "asset_turnover": False,
    "composite_quality_score": False,
}


# ==========================================================
# Database
# ==========================================================


def get_connection():

    return sqlite3.connect(DB_PATH)


def load_data():

    conn = get_connection()

    query = """
    SELECT

        fr.*,

        c.company_name,

        pg.peer_group_name

    FROM financial_ratios fr

    LEFT JOIN companies c
        ON fr.company_id = c.id

    LEFT JOIN peer_groups pg
        ON fr.company_id = pg.company_id
    """

    df = pd.read_sql(query, conn)

    conn.close()

    df = df.dropna(subset=["year"])

    df["year_order"] = df["year"].map(YEAR_PRIORITY)

    df = df.sort_values("year_order").groupby("company_id", as_index=False).last()

    df.drop(columns="year_order", inplace=True)

    return df


# ==========================================================
# Percentile Columns
# ==========================================================


def add_percentile_columns(group):

    group = group.copy()

    for metric, inverse in RANK_METRICS.items():
        if metric not in group.columns:
            continue

        pct = (
            group[metric].rank(
                pct=True,
                ascending=inverse,
                na_option="bottom",
            )
            * 100
        )

        if inverse:
            pct = 100 - pct

        group[f"{metric}_pctile"] = pct.round(1)

    return group


# ==========================================================
# Median Row
# ==========================================================


def add_median_row(group):

    numeric = group.select_dtypes(include="number").median()

    row = {}

    for col in group.columns:
        if col in numeric.index:
            row[col] = round(numeric[col], 2)

        elif col == "company_name":
            row[col] = "Peer Median"

        else:
            row[col] = ""

    return pd.concat(
        [group, pd.DataFrame([row])],
        ignore_index=True,
    )


# ==========================================================
# Workbook Formatting
# ==========================================================


def format_workbook(workbook):

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"

        # Auto Width

        for column_cells in sheet.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0 for cell in column_cells
            )

            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                length + 3, 35
            )

        headers = [cell.value for cell in sheet[1]]

        # Percentile Colours

        for idx, header in enumerate(headers, start=1):
            if not str(header).endswith("_pctile"):
                continue

            for row in range(2, sheet.max_row):
                value = sheet.cell(row, idx).value

                if not isinstance(value, (int, float)):
                    continue

                if value >= 75:
                    sheet.cell(row, idx).fill = GREEN

                elif value >= 25:
                    sheet.cell(row, idx).fill = YELLOW

                else:
                    sheet.cell(row, idx).fill = RED

        # Highlight Best Company

        if "composite_quality_score" in headers:
            score_col = headers.index("composite_quality_score") + 1

            best_row = 2
            best_score = -1

            for r in range(2, sheet.max_row):
                value = sheet.cell(r, score_col).value

                if isinstance(value, (int, float)):
                    if value > best_score:
                        best_score = value
                        best_row = r

            for c in range(1, sheet.max_column + 1):
                sheet.cell(best_row, c).fill = GOLD


# ==========================================================
# Export
# ==========================================================


def export_peer_report():

    df = load_data()

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:
        for peer_group, group in df.groupby("peer_group_name"):
            if pd.isna(peer_group):
                continue

            group = add_percentile_columns(group)

            group = group.sort_values(
                "composite_quality_score",
                ascending=False,
            )

            group = add_median_row(group)

            group.to_excel(
                writer,
                sheet_name=str(peer_group)[:31],
                index=False,
            )

    workbook = load_workbook(OUTPUT_PATH)

    format_workbook(workbook)

    workbook.save(OUTPUT_PATH)

    print("=" * 60)
    print("Peer Comparison Report Generated")
    print("=" * 60)
    print(f"Peer Groups : {df['peer_group_name'].nunique()}")
    print(f"Companies   : {len(df)}")
    print(f"Saved To    : {OUTPUT_PATH}")


# ==========================================================
# Main
# ==========================================================


def main():

    export_peer_report()


if __name__ == "__main__":
    main()
